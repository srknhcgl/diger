def ListExclFiles(directory):
    import os 
    #from os import listdir
    return  os.listdir(directory)
################################################
def GetSheetNames(fileName):
    from openpyxl import load_workbook
    wb2 = load_workbook(directory+"/"+fileName)
    result= []
    for s in wb2.get_sheet_names():        
        if(isInt(s)):            
            result.append(s)
    return result
################################################
def isInt(val):
 try:
     if(type(int(val)) is int):
         return True
     else:
         return False         
 except:
    return False
################################################    
def GetCellValue(fileName,sheetName):
    from openpyxl import load_workbook
    wb3 = load_workbook(directory+"/"+fileName)
    ws2=wb3[sheetName]
    for i in range(16,95):        
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AH'+str(i)].value,201401)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AI'+str(i)].value,201402)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AJ'+str(i)].value,201403)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AK'+str(i)].value,201404)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AL'+str(i)].value,201405)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AM'+str(i)].value,201406)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AN'+str(i)].value,201407)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AO'+str(i)].value,201408)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AP'+str(i)].value,201409)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AQ'+str(i)].value,201410)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AR'+str(i)].value,201411)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AS'+str(i)].value,201412)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AW'+str(i)].value,201301)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AX'+str(i)].value,201302)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AY'+str(i)].value,201303)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['AZ'+str(i)].value,201304)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BA'+str(i)].value,201305)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BB'+str(i)].value,201306)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BC'+str(i)].value,201307)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BD'+str(i)].value,201308)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BE'+str(i)].value,201309)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BF'+str(i)].value,201310)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BG'+str(i)].value,201311)
            InsertValuesSql(sheetName, ws2['AG'+str(i)].value,ws2['BH'+str(i)].value,201312)

################################################
def connect_to_db(SERVER, DATABASE, UID, PWD):    
    cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+SERVER+';DATABASE='+DATABASE+';UID='+UID+';PWD='+PWD)
    cursor = cnxn.cursor()
    return cursor,cnxn

################################################            
def InsertValuesSql(restNo,PLDesc,PLAmount,FileName):
    cursor,cnxn=connect_to_db('','','','')
    query="INSERT INTO [dbo].[tbl_] ([Rest_No],[PL_Desc],[PL_Amount],[DateKey]) VALUES (?, ?, ?, ?)"
    cursor.execute(query, restNo,PLDesc,PLAmount,FileName)
    cnxn.commit()
################################################    
import pyodbc
from sqlalchemy import create_engine
engine = create_engine('mssql+pyodbc://')



directory="C:/"
fileName=""



for f in ListExclFiles(directory):
    for sh in GetSheetNames(f):
        GetCellValue(f,sh)
################################################

